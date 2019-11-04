from datetime import datetime, timedelta
from openpyxl import load_workbook
import SmokeAndSleep as sas


#logLocation = "C:/Users/Nagarjuna/Documents/MedLogTest.xlsm"
#from sympy import true

#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201510.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201608.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_Hospital_201506.xlsm"

#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_test2.xlsm"
#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_ForGampopa.xlsm"

logSet = [
  "C:/Users/Gampopa/Documents/MedLog_201412.xlsm",
  "C:/Users/Gampopa/Documents/MedLog_201510.xlsm",
  "C:/Users/Gampopa/Documents/MedLog_201608.xlsm",
  "C:/Users/Gampopa/Documents/MedLog_201706_ForGampopa.xlsm"
  ]

#### Start ####
def loadRows(rows, key, noteKey=None):
  if noteKey != None:
    print("Extracting: {} with note {}".format(key, noteKey))
    loadedRows = [(row[0].value, row[2].value, row[3].value)
                  for row in rows if str(row[1].value) == key and noteKey.lower() in str(row[3].value).lower()]
    times = [row[0] for row in loadedRows]
  else:
    print("Extracting: ", key)
    loadedRows = [(row[0].value, row[2].value, row[3].value)
                  for row in rows if str(row[1].value) == key]
    times = [row[0] for row in loadedRows]
  return loadedRows, times
#### End ####

#### Start ####
def loadRowsAsDict(rows, key, noteKey=None):
  if noteKey != None:
    print("Extracting: {} with note {}".format(key, noteKey))
    loadedRows = {row[0].value : (row[0].value, row[2].value, row[3].value)
                  for row in rows if str(row[1].value).lower() == key.lower() and noteKey.lower() in str(row[3].value).lower()}
  else:
    print("Extracting: ", key)
    loadedRows = {row[0].value : (row[0].value, row[2].value, row[3].value)
                  for row in rows if str(row[1].value).lower() == key.lower()}
  return loadedRows
#### End ####

#### Start ####
def loadWorkbooks(logSet, sleepKey=sas.sleepKey, wakeKey=sas.wakeKey, smokeKey=sas.smokeKey, potSmokeKey=sas.potSmokeKey):

  sleepTimes = set()
  smokes = set()
  smokeTimes = set()
  potSmokes = set()
  potSmokeTimes = set()

  # we have to first load the data into sets, then put it in sorted lists, because we may have duplicate data between workbooks  
  for log in logSet:
    print("loading from workbook: ", log)
    wb = load_workbook(log, data_only=True)
    ws = wb.active
  #  print(wb.sheetnames)
    print("Row Count: " + str(ws.max_row))
    sleepTimes |= {(row[0].value, row[1].value, row[2].value, row[3].value)
                   for row in ws.rows if row[1].value in sleepKey or row[1].value in wakeKey}

    ### Extract the Smoke times
    print("Extracting Smokes", datetime.now())
    smokesTmp, smokeTimesTmp = loadRows(ws.rows, smokeKey)
    smokes.update(smokesTmp)
    smokeTimes.update(smokeTimesTmp)

    # Extract Smoke logs that note Pot Smoke    
    potSmokesTmp, potSmokesTimesTmp = loadRows(ws.rows, smokeKey, potSmokeKey)
    potSmokes.update(potSmokesTmp)
    potSmokeTimes.update(potSmokesTimesTmp)

    # Extract Sleep/Wake logs that note Pot Smoke
    # todo:  maybe?
    
  sleepTimesSorted = sorted(sleepTimes, key=lambda sleepTimes: sleepTimes[0])

  # Add sleep/wake logs that also log smoke, to the smoke times
  sleepWakeSmokes = [row[0] for row in sleepTimes if smokeKey.lower() in str(row[3]).lower()]
  smokeTimes.update(sleepWakeSmokes)

  print("len(sleepWakeSmokes):", len(sleepWakeSmokes))
  print("len(sleepTimes):", len(sleepTimes))

  smokeTimesSorted = sorted(smokeTimes)
  smokesSorted = sorted(smokes, key=lambda smokes: smokes[0])

  potSmokeTimesSorted = sorted(potSmokeTimes)
  potSmokesSorted = sorted(potSmokes, key=lambda potSmokes: potSmokes[0])

  return sleepTimesSorted, smokeTimesSorted, smokesSorted, potSmokeTimesSorted, potSmokesSorted
#### End ####


#### Start ####
# This works to eliminate duplicate data, but it fails on the
# middle-of-night Pot Smoke logs where I logged a wake/smoke/sleep all at the same time.
# So instead, just don't load duplicate data....
def loadWorkbooksAsDicts(logSet, sleepKey=sas.sleepKey, wakeKey=sas.wakeKey, smokeKey=sas.smokeKey, potSmokeKey=sas.potSmokeKey):

  sleepTimes = {}
  smokes = {}
  potSmokes = {}

  # we have to first load the data into sets, then put it in sorted lists, because we may have duplicate data between workbooks  
  for log in logSet:
    print("loading from workbook: ", log)
    wb = load_workbook(log, data_only=True)
    ws = wb.active
  #  print(wb.sheetnames)
    print("Row Count: " + str(ws.max_row))
    sleepTimesTmp = {row[0].value: (row[0].value, row[1].value, row[2].value, row[3].value)
                   for row in ws.rows if row[1].value in sleepKey or row[1].value in wakeKey}

    sleepTimes = {**sleepTimes, **sleepTimesTmp}

#     ### Extract the Smoke times
    print("Extracting Smokes", datetime.now())
    smokesTmp = loadRowsAsDict(ws.rows, smokeKey)
    smokes = {**smokes, **smokesTmp}

    # Extract Smoke logs that note Pot Smoke    
    potSmokesTmp = loadRowsAsDict(ws.rows, smokeKey, potSmokeKey)
    potSmokes = {**potSmokes, **potSmokesTmp}

  sleepTimesSorted = sorted(sleepTimes.values(), key=lambda sleepTimes: sleepTimes[0])

  smokeTimesSorted = sorted(smokes.keys())
  smokesSorted = sorted(smokes.values(), key=lambda smokes: smokes[0])

  potSmokeTimesSorted = sorted(potSmokesTmp.keys())
  potSmokesSorted = sorted(potSmokes.values(), key=lambda smokes: smokes[0])

  return sleepTimesSorted, smokeTimesSorted, smokesSorted, potSmokeTimesSorted, potSmokesSorted
#### End ####

