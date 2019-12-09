#from datetime import timedelta
from datetime import datetime
#import random
from openpyxl import load_workbook
import matplotlib
import matplotlib.pyplot as plt
#import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter

import os

#logLocation = "C:/Users/Nagarjuna/Documents/MedLogTest.xlsm"
#from sympy import true

#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201510.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201608.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_Hospital_201506.xlsm"

#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706.xlsm"
logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_test2.xlsm"


# Parameters
drugHalfLife = 12 * 60 * 60  # seconds
drugDigestionHalfLife = 0.75 * 60 * 60  # seconds
projectionTime = 48 * 60 * 60  # seconds
mvAvgDuration = drugHalfLife * 5
sleepAvgDuration = 2 * 24 * 60 * 60  # seconds
timeStep = 180  # seconds
timeRange = 7 * 24 * 60 * 60  # seconds


# Event keys from log
#medKey = "Trileptal"
medKey = "Lamictal"
sleepKey = ("Sleep (crosseyed)", "Sleep (med-tired)", "Sleep (normal)")
wakePoor = "Wake (poor)"
wakeMedium = "Wake (medium)"
wakeSound = "Wake (sound)"
#wakeKey = ("Wake (medium)", "Wake (poor)", "Wake (sound)")
wakeKey = (wakePoor, wakeMedium, wakeSound)
seizureKey = "Seizure"
jackOffKey = "Jack Off"
showJackOffs = True

#######################################################################################################################
## Functions
#######################################################################################################################

#### Start ####
## The amount that has been digested, less the amount that has been metabolized,
## at a given time (sec) after the last dose.
def concAtTime(concentrationZero, remainingDose, doseValue, sec):
  return (concentrationZero + (remainingDose + doseValue) * (1 - pow(2, - sec / drugDigestionHalfLife))) * \
         pow(2, - sec / drugHalfLife)
#### End ####

#### Start ####
## The amount still to be digested, from all previous doses.
def remainingDoseAtTime(remainingDose, doseValue, sec):
  return (remainingDose + doseValue) * pow(2, - sec / drugDigestionHalfLife)
#### End ####

#######################################################################################################################
## End Functions
#######################################################################################################################

print(matplotlib.matplotlib_fname())


#wb = load_workbook(logLocation, guess_types=True, data_only=True)
wb = load_workbook(logLocation, data_only=True)
#print(wb.get_sheet_names())
ws = wb.active

rowCount = ws.max_row
print("Row Count: " + str(rowCount))

#for row in ws.get_squared_range(1, 1, 3, rowCount):
#  print(row[0].value, row[1].value, row[2].value)
#  if row[1].value == key:
#    print(row[0].value, row[1].value, row[2].value)

# Extract the Doses
doses = [(row[0].value, row[2].value)
         for row in ws.rows if row[1].value == medKey]
#for row in ws.get_squared_range(1, 2, 3, rowCount) if row[1].value == medKey]
#print(doses)
#for row in doses:
#  print(row[0], row[1])

doseTimes = [row[0] for row in doses]
doseValues = [row[1] for row in doses]

curve = list()
curveTimes = list()
curveConcs = list()
concentration = 0
timeZero = 0
concentrationZero = 0
remainingDose = 0
for doseIdx in range(0, len(doseTimes) - 1):
  #for doseIdx in range(0, 15):
  timeZero = doseTimes[doseIdx].timestamp()
  #  curve.append([times[doseIdx], concentrationZero])
  #  print("T:", times[doseIdx + 1].timestamp() - timeZero)
  for sec in range(0, int((doseTimes[doseIdx + 1] - doseTimes[doseIdx]).total_seconds()), timeStep):
    time = timeZero + sec
    concentration = concAtTime(concentrationZero, remainingDose, doseValues[doseIdx], sec)
    curveTimes.append(datetime.fromtimestamp(time))
    curveConcs.append(concentration)
  # curve.append([datetime.fromtimestamp(time), concentration])
#  print("remaining before: ", remainingDose, " doseIdx: ", doseIdx, " doseValue: ", doseValues[doseIdx])
  finalIntervalTime = (doseTimes[doseIdx + 1] - doseTimes[doseIdx]).total_seconds()
#  finalIntervalTime = (doseTimes[-1] - doseTimes[-2]).total_seconds()
  concentrationZero = concAtTime(concentrationZero, remainingDose, doseValues[doseIdx], finalIntervalTime)
  remainingDose = remainingDoseAtTime(remainingDose, doseValues[doseIdx], finalIntervalTime)
#  print("remaining after: ", remainingDose)
#  print("outside sec: ", sec, " in min: ", sec / 60, " in hr: ", sec / 3600)

# Project into the future
projCurveTimes = list()
projCurveConcs = list()
timeZero = doseTimes[-1].timestamp()
for sec in range(0, projectionTime, timeStep):
  time = timeZero + sec
  concentration = concAtTime(concentrationZero, remainingDose, doseValues[-1], sec)
  projCurveTimes.append(datetime.fromtimestamp(time))
  projCurveConcs.append(concentration)
#  curve.append([datetime.fromtimestamp(time), concentration])

# Find moving averages of to-date data
fullDurationTicks = int(mvAvgDuration / timeStep)
mvAvgTime = curveTimes[0]
mvAvgConc = curveConcs[0]
mvAvgCurveTimes = list()
mvAvgCurveConcs = list()
mvAvgCurveTimes.append(mvAvgTime)
mvAvgCurveConcs.append(mvAvgConc)
for i in range(1, len(curveTimes)):
  mvAvgTime = curveTimes[i]
  if i < fullDurationTicks:
    mvAvgConc = (mvAvgConc * i + curveConcs[i]) / (i + 1)
  else:
    mvAvgConc = (mvAvgConc * fullDurationTicks - curveConcs[i - fullDurationTicks] + curveConcs[i]) / fullDurationTicks
  mvAvgCurveTimes.append(mvAvgTime)
  mvAvgCurveConcs.append(mvAvgConc)

# Project moving average into future
# note: assume that we have at least fullDurationTicks of real data, previously calculated
for i in range(0, len(projCurveTimes)):
  mvAvgTime = projCurveTimes[i]
  if i < fullDurationTicks:
    mvAvgConc = (mvAvgConc * fullDurationTicks - curveConcs[-(fullDurationTicks - i)] + projCurveConcs[i]) / fullDurationTicks
  else:
    mvAvgConc = (mvAvgConc * fullDurationTicks - projCurveConcs[i - fullDurationTicks] + projCurveConcs[i]) / fullDurationTicks
  mvAvgCurveTimes.append(mvAvgTime)
  mvAvgCurveConcs.append(mvAvgConc)

#print("", ws.get_squared_range(1, 1, 3, rowCount))

# Extract Sleep Times
sleepTimes = [(row[0].value, row[1].value)
              for row in ws.rows if row[1].value in sleepKey or row[1].value in wakeKey]
#for row in ws.get_squared_range(1, 1, 3, rowCount) if row[1].value in sleepKey or row[1].value in wakeKey]
#print("STimes:", sleepTimes)
##for row in sleepTimes:
#for i in range(0, len(sleepTimes), 2):
#  print("S:", sleepTimes[i], sleepTimes[i + 1])
#  print("TD:", round((sleepTimes[i + 1][0] - sleepTimes[i][0]).seconds / 3600, 2))

# Calculate moving averages of sleep times
sleepDurationTicks = int(sleepAvgDuration / timeStep)
sleepDurations = list()
#sleepDurations.append(0)
lastStartIdx = 0
lastEndIdx = 0
sleepDur = 0
for i in range(0, len(mvAvgCurveTimes)):
  # print("lastEndIdx: ", lastEndIdx)
  # print("len:", len(sleepTimes))
  # print("mvAvgCurveTemis[i]", mvAvgCurveTimes[i])
  # print("sleepTimes[lastEndIdx + 1][0]", sleepTimes[lastEndIdx + 1][0])

  if lastEndIdx + 1 < len(sleepTimes) and mvAvgCurveTimes[i] > sleepTimes[lastEndIdx + 1][0]:
    lastEndIdx += 1
  if lastStartIdx + 1 < len(sleepTimes) and \
          i - sleepDurationTicks >= 0 and \
          mvAvgCurveTimes[i - sleepDurationTicks] > sleepTimes[lastStartIdx + 1][0]:
    lastStartIdx += 1
  # If the last index is even, we're in a sleep period
  if lastEndIdx / 2.0 == int(lastEndIdx / 2):
    sleepDur += timeStep
  if i >= sleepDurationTicks:
    if lastStartIdx / 2.0 == int(lastStartIdx / 2):
      sleepDur -= timeStep
  sleepDurations.append(24 * sleepDur / sleepAvgDuration)
#  mvAvgSleepDurations.append(random.randint(0, 24))

# Extract seizures
seizureTimes = [row[0].value
                for row in ws.rows if row[1].value == seizureKey]
#for time in seizureTimes:
#  print(time)

# Extract Jack Offs
jackOffTimes = [row[0].value
                for row in ws.rows if row[1].value == jackOffKey]

#####################
## Do the plotting ##
#####################


#fig = plt.gcf()
#fig = plt.figure()
#dpi = fig.get_dpi()
#defSize = fig.get_size_inches()
#print("dpi: ", dpi)
#print("size: ", defSize)
## The figure() fn has to be called before anything else, apparently....
## default dpi: 80
## default size [8, 6]
#fig.set_size_inches(defSize[0], 1920 / dpi)

fig = plt.figure(num=1, figsize=((1920 - 5) / 80, 10))
plt.subplots_adjust(left=0.02, right=0.98)
## For printing
#fig = plt.figure(num=1, figsize=((1920 - 5) / 80 / 2.3, 10))
#plt.subplots_adjust(left=0.05, right=0.96)

#fig.canvas.manager.window.move(0,0)


matplotlib.use("TkAgg")

# http://stackoverflow.com/questions/7449585/how-do-you-set-the-absolute-position-of-figure-windows-with-matplotlib
mng = plt.get_current_fig_manager()
mng.window.wm_geometry("+0+100")  # TkAgg
#mng.window.SetPosition((500, 0)) # wx

#Do the plotting

plt.plot(doseTimes, doseValues, 'b+', markersize=10, markeredgewidth=2)
### turn off plotting of drug curves
#plt.plot(curveTimes, curveConcs)
#plt.plot(projCurveTimes, projCurveConcs, '#444444')
#plt.plot(mvAvgCurveTimes, mvAvgCurveConcs, 'c')

for seizureTime in seizureTimes:
  plt.axvline(seizureTime, color="r")

if showJackOffs:
  for jackOffTime in jackOffTimes:
    plt.axvline(jackOffTime, color="y")

# Now
plt.axvline(datetime.now(), color="k")

# Set the time range to plot
## Python doesn't have a switch/case?? Intentionally so?  https://www.python.org/dev/peps/pep-3103/
#lastWeek = True
lastWeek = True
plotStartTime = curveTimes[0]
if lastWeek:
  plotStartTime = datetime.fromtimestamp(datetime.now().timestamp() - timeRange)

# 2015-06-22  15:05:00 -- last medicine dose before hospital
# todo:  add end datetime other than "now"
#plotStartTime = datetime(2015, 6, 22, 15, 5, 0, 0)
#projectionTime = 0

startIdx = 0
for i in range(0, len(curveTimes)):
  if curveTimes[i] > plotStartTime:
    startIdx = i
    break
maxConc = 0
for conc in curveConcs[startIdx:]:
  if conc > maxConc:
    maxConc = conc
for conc in projCurveConcs:
  if conc > maxConc:
    maxConc = conc
doseStartIdx = 0
for i in range(0, len(doseTimes)):
  if doseTimes[i] > plotStartTime:
    doseStartIdx = i
    break
for dose in doseValues[doseStartIdx:]:
  if dose > maxConc:
    maxConc = dose

maxLim = int(maxConc * 1.05 / 100 + 0.5) * 100 + 50
plt.ylim([0, maxLim])
#print("maxLim ==", maxLim)
#print("plt.ylim()[1]", plt.ylim()[1])

# Sleep/wake intervals
for i in range(0, len(sleepTimes), 2):
  color = ()
  if len(sleepTimes) <= i + 1:
    print("Probably forgot to log a \"wake\"")
    exit(1)
  if wakePoor == sleepTimes[i + 1][1]:
    color = (0.25, 0.25, 0.25, 0.20)
  elif wakeMedium == sleepTimes[i + 1][1]:
    color = (0.25, 0.25, 0.25, 0.40)
  elif wakeSound == sleepTimes[i + 1][1]:
    color = (0.25, 0.25, 0.25, 0.65)
  plt.fill_between([sleepTimes[i][0], sleepTimes[i + 1][0]], plt.ylim()[1], color=color)
##  plt.xticks(list(plt.xticks()[0]) + [sleepTimes[i].value, sleepTimes[i + 1].value])
##  plt.text(sleepTimes[i], 0, sleepTimes[i], fontsize="small", horizontalalignment="center", verticalalignment="top")
  plt.text(sleepTimes[i][0], 0, sleepTimes[i][0].strftime("%m-%d %H:%M"),
           fontdict={"size": "small"}, horizontalalignment="center", verticalalignment="center", rotation=60)
  if (i < len(sleepTimes) - 1):
    plt.text(sleepTimes[i + 1][0], maxLim, sleepTimes[i + 1][0].strftime("%m-%d %H:%M"),
             fontsize="small", horizontalalignment="center", verticalalignment="center", rotation=60)
    sleepDuration = round((sleepTimes[i + 1][0] - sleepTimes[i][0]).total_seconds() / 3600, 1)
    loc = sleepTimes[i][0] + (sleepTimes[i + 1][0] - sleepTimes[i][0])/2
    plt.text(loc, plt.ylim()[1]/8, sleepDuration,
             fontsize="medium", horizontalalignment="center", verticalalignment="center")
  if (i < len(sleepTimes) - 2):
    wakeDuration = round((sleepTimes[i + 2][0] - sleepTimes[i + 1][0]).total_seconds() / 3600, 1)
    loc = sleepTimes[i + 1][0] + (sleepTimes[i + 2][0] - sleepTimes[i + 1][0])/2
    plt.text(loc, plt.ylim()[1]/8, wakeDuration,
             fontsize="medium", horizontalalignment="center", verticalalignment="center")

sleepDuration = round((datetime.now() - sleepTimes[-1][0]).total_seconds() / 3600, 1)
loc = sleepTimes[-1][0] + (datetime.now() - sleepTimes[-1][0])/2
plt.text(loc, plt.ylim()[1]/8, sleepDuration,
         fontsize="medium", horizontalalignment="center", verticalalignment="center")

# Intervals between medicine doses
yloc = (maxLim - 300) / 2 + 300
#yloc = 7 * plt.ylim()[1] / 8
for i in range(0, len(doseTimes) - 1):
  doseInterval = round((doseTimes[i + 1] - doseTimes[i]).total_seconds() / 3600, 1)
  loc = doseTimes[i] + (doseTimes[i + 1] - doseTimes[i])/2
  plt.text(loc, yloc, doseInterval, fontsize="medium", horizontalalignment="center", verticalalignment="center", color='b')

doseInterval = round((datetime.now() - doseTimes[-1]).total_seconds() / 3600, 1)
loc = doseTimes[-1] + (datetime.now() - doseTimes[-1])/2
plt.text(loc, yloc, doseInterval, fontsize="medium", horizontalalignment="center", verticalalignment="center", color='b')

###

# Fix up the axes
ax = plt.gca()

# Create an axis for sleep times
axSleep = ax.twinx()
axSleep.plot(mvAvgCurveTimes[:len(sleepDurations)], sleepDurations, "k--")
axSleep.set_ylim(0, 24)
sleepMajorLocator = MultipleLocator(4)
sleepMinorLocator = MultipleLocator(2)
axSleep.yaxis.set_major_locator(sleepMajorLocator)
axSleep.yaxis.set_minor_locator(sleepMinorLocator)
#axSleep.set_yticks([0, 8, 16, 24])
#axSleep.grid(b=True, which=u'both', axis='y', color='k', linestyle='--')
axSleep.grid(b=True, which=u'both', axis='y', color='k')


plt.xlim(xmin=plotStartTime)
#plt.ylim(ymax=650)

# Set the max dose axis value
ax.set_ylim(ymax=maxLim)
#plt.minorticks_on()

ax.relim()
ax.autoscale_view(True, True, True)
#ax.tick_params(axis='x', which='minor', bottom='off')
#ax.tick_params(axis='y', which='minor', bottom='on')
ax.grid(b=True, which=u'both', color='b', linestyle='-.')


plt.show()



