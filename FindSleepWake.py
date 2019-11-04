#from datetime import datetime, timedelta
from openpyxl import load_workbook

from pandas.plotting import register_matplotlib_converters
register_matplotlib_converters()

###
# Utility to find the row of missing "wake" logs
###

#logLocation = "C:/Users/Nagarjuna/Documents/MedLogTest.xlsm"
#from sympy import true

#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201510.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201608.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_Hospital_201506.xlsm"

#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_test2.xlsm"
logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_ForGampopa.xlsm"


# Event keys from log
sleepKey = ("Sleep (crosseyed)", "Sleep (med-tired)", "Sleep (normal)")
wakePoor = "Wake (poor)"
wakeMedium = "Wake (medium)"
wakeSound = "Wake (sound)"
#wakeKey = ("Wake (medium)", "Wake (poor)", "Wake (sound)")
wakeKey = (wakePoor, wakeMedium, wakeSound)


#wb = load_workbook(logLocation, guess_types=True, data_only=True)
wb = load_workbook(logLocation, data_only=True)
#print(wb.get_sheet_names())
ws = wb.active

rowNum = 0
expectWake = False
for row in ws.rows:
  if row[1].value in sleepKey:
    expectWake = True
  elif expectWake == True:
      if row[1].value not in wakeKey:
        print("no wake on row: ", rowNum)
      expectWake = False
  rowNum += 1
  


                  
              
              






