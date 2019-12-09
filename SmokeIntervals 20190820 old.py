from datetime import datetime, timedelta
from openpyxl import load_workbook
import matplotlib
import matplotlib.pyplot as plt
from matplotlib import rc, rcParams
import matplotlib.dates as dates

import WorkbookLoaders as wl

from pandas.plotting import register_matplotlib_converters
register_matplotlib_converters()

print(matplotlib.matplotlib_fname())

#import os

#######################################################################################################################
## Functions
#######################################################################################################################

#### Start ####
def findTimeDiffs(rows, startTime, endTime):
  totalCount = 0
  timediffs = list()
  for i in range(0, len(smokes) - 1):
  #  print("row 1: ", smokes[i])
  #  print ("row 2: ", smokes[i+1])
    if rows[i][0] < startTime or rows[i][0] > endTime:
      continue
    timediff = rows[i + 1][0] - rows[i][0]
  #  print("timeOne - timeTwo == timediff: {} - {} == {}".format(smokes[i+1][0], smokes[i][0], timediff))
    if timediff < timedelta(0, hour_limit * 60 * 60):
      timediffs.append(timediff.total_seconds() / 60)
    totalCount += 1
  return timediffs, totalCount
#### End ####


#### Start ####
def makeHistogram(timediffs, totalCount, startTime, endTime, hour_limit):
  timePeriodDays = int((endTime - startTime).total_seconds() / (60 * 60 * 24))
  numBins = int(hour_limit * 60 / 3)
  n, bins, patches = plt.hist(timediffs, numBins, range=(0, hour_limit * 60),
                              facecolor='red', alpha=0.5,
                              edgecolor='k', linewidth=0.5,
                              label="Total Fixes: %d (%.1f/day)" % (totalCount, totalCount/timePeriodDays))
  #n, bins, patches = plt.hist(timediffs, numBins, facecolor='red', alpha=0.5, label=r"Time Between Fixes: $$\bf%s$$ to $$\bf%s$$" % (startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")))
  
  print("bin size: ", bins[1] - bins[0])
  
  plt.xlim(xmin=0, xmax=bins[-1])
  
  ax = plt.gca()
  ax.grid(b=True, which=u'both', axis="x", color='k', linestyle='-.')
  
  plt.suptitle("Times Between Nicotine/Cigarette Fixes", fontsize=16)
  plt.title(r"For the %s day period: $\bf%s$ to $\bf%s$" % (timePeriodDays,
             startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")))
  
  plt.xlabel("Time Between Fixes (minutes)")
  plt.ylabel("Count")
  
  plt.legend()
  
  plt.savefig("Fixes %s to %s.png" % (startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")))
  
  plt.show()
#### End ####

#######################################################################################################################
## End Functions
#######################################################################################################################







#logLocation = "C:/Users/Nagarjuna/Documents/MedLogTest.xlsm"
#from sympy import true

#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201510.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201608.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Nagarjuna/Documents/MedLog_Hospital_201506.xlsm"

#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706.xlsm"
#logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_test2.xlsm"
logLocation = "C:/Users/Gampopa/Documents/MedLog_201706_ForGampopa.xlsm"

###
## Parameters
###

###
## Define when the graph will start and end
###
dayRange = 90  # days
currTime = datetime.now()
startTime = currTime - timedelta(dayRange)
endTime = currTime

# don't add time diffs if the difference is larger than this in hours
hour_limit = 2
# key to identify smoke log lines
smokeKey = "Smoke"

#startTime = datetime.min # all data

#startTime = datetime.strptime("2019-01-01 00:00:00", "%Y-%m-%d %H:%M:%S")
#endTime = datetime.strptime("2019-01-31 00:00:00", "%Y-%m-%d %H:%M:%S")
#endTime = datetime.now()
#currTime = endTime

print("startTime: ", startTime)
print("endTime: ", endTime)
print("currTime: ", currTime)


#wb = load_workbook(logLocation, guess_types=True, data_only=True)
wb = load_workbook(logLocation, data_only=True)
#print(wb.get_sheet_names())
ws = wb.active

rowCount = ws.max_row
print("Row Count: " + str(rowCount))


### Extract the Smoke times
print("Extracting Smokes", datetime.now())
#smokes = [(row[0].value, row[2].value, row[3].value)
#          for row in ws.rows if row[1].value == smokeKey]
smokes, smokeTimes = wl.loadRows(ws.rows, smokeKey)
#smokeTimes = [row[0] for row in smokes]
print("len(smokeTimes): ", len(smokeTimes))

print("Done Extracting", datetime.now())



# totalCount = 0
# timediffs = list()
# for i in range(0, len(smokes) - 1):
# #  print("row 1: ", smokes[i])
# #  print ("row 2: ", smokes[i+1])
#   if smokes[i][0] < startTime or smokes[i][0] > endTime:
#     continue
#   timediff = smokes[i + 1][0] - smokes[i][0]
# #  print("timeOne - timeTwo == timediff: {} - {} == {}".format(smokes[i+1][0], smokes[i][0], timediff))
#   if timediff < timedelta(0, hour_limit * 60 * 60):
#     timediffs.append(timediff.total_seconds() / 60)
#   totalCount += 1

timediffs, totalCount = findTimeDiffs(smokes, startTime, endTime)

print("len(timediffs): ", len(timediffs))    
print("totalCount: ", totalCount)    

makeHistogram(timediffs, totalCount, startTime, endTime, hour_limit)


#xAxLabel = r"Time Between Fixes: $$\bf%s$$ to $$\bf%s$$" % (startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")) 

# timePeriodDays = int((endTime - startTime).total_seconds() / (60 * 60 * 24))
# numBins = int(hour_limit * 60 / 3)
# n, bins, patches = plt.hist(timediffs, numBins, range=(0, hour_limit * 60),
#                             facecolor='red', alpha=0.5,
#                             edgecolor='k', linewidth=0.5,
#                             label="Total Fixes: %d (%.1f/day)" % (totalCount, totalCount/timePeriodDays))
# #n, bins, patches = plt.hist(timediffs, numBins, facecolor='red', alpha=0.5, label=r"Time Between Fixes: $$\bf%s$$ to $$\bf%s$$" % (startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")))
# 
# print("bin size: ", bins[1] - bins[0])
# 
# plt.xlim(xmin=0, xmax=bins[-1])
# 
# # activate latex text rendering
# #rc('text', usetex=True)
# #rc('axes', linewidth=2)
# #rc('font', weight='bold')
# #rcParams['text.latex.preamble'] = [r'\usepackage{sfmath} \boldmath']
# 
# 
# ax = plt.gca()
# ax.grid(b=True, which=u'both', axis="x", color='k', linestyle='-.')
# 
# 
# 
# plt.suptitle("Times Between Nicotine/Cigarette Fixes", fontsize=16)
# plt.title(r"For the %s day period: $\bf%s$ to $\bf%s$" % (timePeriodDays,
#            startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")))
# 
# plt.xlabel("Time Between Fixes (minutes)")
# plt.ylabel("Count")
# 
# plt.legend()
# 
# plt.savefig("Fixes %s to %s.png" % (startTime.strftime("%Y-%m-%d"), endTime.strftime("%Y-%m-%d")))
# 
# plt.show()











































